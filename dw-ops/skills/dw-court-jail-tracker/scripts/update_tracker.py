#!/usr/bin/env python3
"""
Merge scraped JusticeWorks case rows into the firm Court & Jail Visit Tracker xlsx.

Inputs:
  --scraped   Path to JSON file produced by the Chrome scrape step. Schema:
              [
                {
                  "docket": "16090-21",
                  "section": "E",
                  "ada": "Hall Lea R.",
                  "client_name": "ANTONIO DEVON HADNOT",
                  "charges": "First Degree Murder (2 Counts)",
                  "next_court_event": "Trial Date",
                  "next_court_date": "2026-04-13",      # ISO; we'll re-format
                  "trial_date": "2026-04-13",            # may be ""
                  "in_custody": true,
                  "last_jail_visit": "2026-03-10"        # may be "" — only used if tracker JAIL VISIT is blank
                },
                ...
              ]
  --tracker   Path to the Court & Jail Visit Tracker.xlsx

Behavior:
  - Backs up the tracker to ~/.dw-tracker/backups/Tracker-YYYY-MM-DD.xlsx.
  - Updates existing rows matched by DOCKET # (does NOT touch JAIL VISIT).
  - Appends new rows.
  - Recomputes JAIL VISIT NEEDED for in-custody cases.
  - Prints a JSON summary to stdout.

Output (stdout): JSON {"new_cases":[...], "updated_cases":[...], "overdue_visits":[...]}
"""

from __future__ import annotations

import argparse
import json
import shutil
import sys
from datetime import date, datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl not installed. Install with: pip install openpyxl --break-system-packages",
          file=sys.stderr)
    sys.exit(1)

OVERDUE_DAYS = 30

# Column index (1-based, openpyxl style) for each header.
# Built from the existing tracker layout:
# A DOCKET # | B SECTION | C ADA | D CLIENT NAME | E CHARGES |
# F NEXT COURT EVENT | G NEXT COURT DATE | H TRIAL DATE | I JAIL VISIT | J JAIL VISIT NEEDED
COLS = {
    "docket": 1, "section": 2, "ada": 3, "client_name": 4, "charges": 5,
    "next_court_event": 6, "next_court_date": 7, "trial_date": 8,
    "jail_visit": 9, "jail_visit_needed": 10,
}


def parse_date(s):
    """Accept ISO ('2026-04-13'), US ('04/13/2026'), or empty/None. Return date or None."""
    if not s or not str(s).strip():
        return None
    s = str(s).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    # openpyxl may have read a real date object
    if isinstance(s, (date, datetime)):
        return s.date() if isinstance(s, datetime) else s
    return None


def fmt_date(d):
    """Tracker display format: MM/DD/YYYY."""
    if d is None:
        return ""
    if isinstance(d, str):
        d = parse_date(d)
        if d is None:
            return ""
    return d.strftime("%m/%d/%Y")


def cell_to_date(cell_value):
    """openpyxl cell value → date (handles datetime, str, or None)."""
    if cell_value is None:
        return None
    if isinstance(cell_value, datetime):
        return cell_value.date()
    if isinstance(cell_value, date):
        return cell_value
    return parse_date(str(cell_value))


def compute_visit_status(last_visit_date, today):
    """Return the JAIL VISIT NEEDED string."""
    if last_visit_date is None:
        return "NEVER VISITED"
    days = (today - last_visit_date).days
    if days > OVERDUE_DAYS:
        return f"OVERDUE ({days} days)"
    return f"OK ({days} days)"


def backup_tracker(tracker_path: Path) -> Path:
    backup_dir = Path.home() / ".dw-tracker" / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    dest = backup_dir / f"Tracker-{date.today().isoformat()}.xlsx"
    shutil.copy2(tracker_path, dest)
    return dest


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--scraped", required=True, type=Path)
    ap.add_argument("--tracker", required=True, type=Path)
    ap.add_argument("--today", default=None,
                    help="Override today's date (YYYY-MM-DD) for testing")
    args = ap.parse_args()

    today = parse_date(args.today) if args.today else date.today()

    if not args.tracker.exists():
        print(f"Tracker not found: {args.tracker}", file=sys.stderr)
        sys.exit(2)
    if not args.scraped.exists():
        print(f"Scraped file not found: {args.scraped}", file=sys.stderr)
        sys.exit(2)

    scraped = json.loads(args.scraped.read_text())
    if not isinstance(scraped, list):
        print("Scraped JSON must be a list of case objects", file=sys.stderr)
        sys.exit(2)

    backup_path = backup_tracker(args.tracker)

    wb = load_workbook(args.tracker)
    ws = wb.active

    # ws.max_row counts formatted-but-empty rows (the tracker template has 998).
    # Find the real last row by scanning for the last non-empty docket.
    real_last_row = 1
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=COLS["docket"]).value not in (None, ""):
            real_last_row = row_idx

    # Index existing rows by docket (string, normalized)
    existing = {}
    for row_idx in range(2, real_last_row + 1):
        docket = ws.cell(row=row_idx, column=COLS["docket"]).value
        if docket is None:
            continue
        existing[str(docket).strip()] = row_idx

    new_cases = []
    updated_cases = []
    overdue_visits = []

    for case in scraped:
        docket = str(case.get("docket", "")).strip()
        if not docket:
            continue

        if docket in existing:
            row = existing[docket]
            # Update everything EXCEPT jail_visit (column I)
            for field in ("section", "ada", "client_name", "charges", "next_court_event"):
                ws.cell(row=row, column=COLS[field], value=case.get(field, "") or "")
            ws.cell(row=row, column=COLS["next_court_date"],
                    value=fmt_date(case.get("next_court_date")))
            ws.cell(row=row, column=COLS["trial_date"],
                    value=fmt_date(case.get("trial_date")))
            updated_cases.append({"docket": docket, "client": case.get("client_name", "")})
        else:
            real_last_row += 1
            row = real_last_row
            ws.cell(row=row, column=COLS["docket"], value=docket)
            ws.cell(row=row, column=COLS["section"], value=case.get("section", ""))
            ws.cell(row=row, column=COLS["ada"], value=case.get("ada", ""))
            ws.cell(row=row, column=COLS["client_name"], value=case.get("client_name", ""))
            ws.cell(row=row, column=COLS["charges"], value=case.get("charges", ""))
            ws.cell(row=row, column=COLS["next_court_event"], value=case.get("next_court_event", ""))
            ws.cell(row=row, column=COLS["next_court_date"], value=fmt_date(case.get("next_court_date")))
            ws.cell(row=row, column=COLS["trial_date"], value=fmt_date(case.get("trial_date")))
            # JAIL VISIT stays blank for new cases — Chris fills it on first visit
            ws.cell(row=row, column=COLS["jail_visit"], value="")
            existing[docket] = row
            new_cases.append({
                "docket": docket,
                "client": case.get("client_name", ""),
                "in_custody": bool(case.get("in_custody")),
            })

        # Recompute JAIL VISIT NEEDED for in-custody cases
        if case.get("in_custody"):
            jail_visit_cell = ws.cell(row=existing[docket], column=COLS["jail_visit"]).value
            last_visit = cell_to_date(jail_visit_cell)
            status = compute_visit_status(last_visit, today)
            ws.cell(row=existing[docket], column=COLS["jail_visit_needed"], value=status)
            if status.startswith("OVERDUE") or status == "NEVER VISITED":
                overdue_visits.append({
                    "docket": docket,
                    "client": case.get("client_name", ""),
                    "status": status,
                    "charges": case.get("charges", ""),
                    "next_court_date": fmt_date(case.get("next_court_date")),
                })
        else:
            # Out of custody — no jail visit needed
            ws.cell(row=existing[docket], column=COLS["jail_visit_needed"], value="N/A (out)")

    wb.save(args.tracker)

    summary = {
        "backup_path": str(backup_path),
        "tracker_path": str(args.tracker),
        "today": today.isoformat(),
        "new_cases": new_cases,
        "updated_cases": updated_cases,
        "overdue_visits": overdue_visits,
    }
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
