#!/usr/bin/env python3
"""
Build the "who Chris needs to see this week" list from the updated tracker.

Reads the tracker xlsx and produces four buckets:
  - jail_visits_overdue
  - court_within_7_days
  - trial_within_30_days
  - new_cases (since the previous run, tracked in ~/.dw-tracker/last-run.json)

Output: JSON to stdout.
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl not installed. Install with: pip install openpyxl --break-system-packages",
          file=sys.stderr)
    sys.exit(1)

COURT_SOON_DAYS = 7
TRIAL_SOON_DAYS = 30

COLS = {
    "docket": 1, "section": 2, "ada": 3, "client_name": 4, "charges": 5,
    "next_court_event": 6, "next_court_date": 7, "trial_date": 8,
    "jail_visit": 9, "jail_visit_needed": 10,
}


def parse_date(s):
    if s is None or s == "":
        return None
    if isinstance(s, datetime):
        return s.date()
    if isinstance(s, date):
        return s
    s = str(s).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--tracker", required=True, type=Path)
    ap.add_argument("--update-summary", required=True, type=Path,
                    help="JSON output from update_tracker.py (used for new-case detection)")
    ap.add_argument("--today", default=None,
                    help="Override today's date (YYYY-MM-DD) for testing")
    ap.add_argument("--last-run-file", default=str(Path.home() / ".dw-tracker" / "last-run.json"),
                    help="Path to the last-run state file")
    args = ap.parse_args()

    today = parse_date(args.today) if args.today else date.today()
    court_cutoff = today + timedelta(days=COURT_SOON_DAYS)
    trial_cutoff = today + timedelta(days=TRIAL_SOON_DAYS)

    update_summary = json.loads(args.update_summary.read_text())

    wb = load_workbook(args.tracker, data_only=True)
    ws = wb.active

    # Find real last row (template has formatted blanks beyond the real data)
    real_last_row = 1
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=COLS["docket"]).value not in (None, ""):
            real_last_row = row_idx

    jail_visits_overdue = []
    court_within_7_days = []
    trial_within_30_days = []

    for row_idx in range(2, real_last_row + 1):
        docket = ws.cell(row=row_idx, column=COLS["docket"]).value
        if docket is None or docket == "":
            continue
        client = ws.cell(row=row_idx, column=COLS["client_name"]).value or ""
        charges = ws.cell(row=row_idx, column=COLS["charges"]).value or ""
        next_event = ws.cell(row=row_idx, column=COLS["next_court_event"]).value or ""
        next_court = parse_date(ws.cell(row=row_idx, column=COLS["next_court_date"]).value)
        trial = parse_date(ws.cell(row=row_idx, column=COLS["trial_date"]).value)
        visit_needed = (ws.cell(row=row_idx, column=COLS["jail_visit_needed"]).value or "")

        base = {
            "docket": str(docket).strip(),
            "client": str(client),
            "charges": str(charges),
        }

        if "OVERDUE" in str(visit_needed) or "NEVER VISITED" in str(visit_needed):
            jail_visits_overdue.append({**base, "status": str(visit_needed)})

        if next_court and today <= next_court <= court_cutoff:
            court_within_7_days.append({**base, "when": next_court.strftime("%m/%d/%Y"),
                                        "event": str(next_event)})

        if trial and today <= trial <= trial_cutoff:
            trial_within_30_days.append({**base, "when": trial.strftime("%m/%d/%Y")})

    # New cases: from update_summary (definitive — they were added this run)
    new_cases = update_summary.get("new_cases", [])

    # Persist last-run state (mostly for change detection across weeks)
    last_run_path = Path(args.last_run_file)
    last_run_path.parent.mkdir(parents=True, exist_ok=True)
    last_run_path.write_text(json.dumps({
        "ran_at": today.isoformat(),
        "case_count": real_last_row - 1,
    }, indent=2))

    out = {
        "today": today.isoformat(),
        "jail_visits_overdue": jail_visits_overdue,
        "court_within_7_days": sorted(court_within_7_days, key=lambda x: x["when"]),
        "trial_within_30_days": sorted(trial_within_30_days, key=lambda x: x["when"]),
        "new_cases": new_cases,
        "counts": {
            "jail_visits_overdue": len(jail_visits_overdue),
            "court_within_7_days": len(court_within_7_days),
            "trial_within_30_days": len(trial_within_30_days),
            "new_cases": len(new_cases),
        },
    }
    print(json.dumps(out, indent=2))


if __name__ == "__main__":
    main()
